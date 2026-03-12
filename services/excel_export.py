"""
Excel Export: fills the Prebid V31 Template with company data.
Only INPUT cells are written — all 90+ formulas remain intact for Excel to recalculate.
"""
import os
import io
import logging
import datetime
import openpyxl
from openpyxl.styles import Protection

logger = logging.getLogger('excel_export')

TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(__file__)),
    'Prebid V31  Template.xlsx',
)


def _f(val, default=0.0):
    """Safe float conversion — always returns a float (uses default when missing)."""
    try:
        v = float(val)
        return v if v == v else default  # NaN check
    except (TypeError, ValueError):
        return default


def _fopt(val):
    """Optional float — returns None for missing/empty values, float otherwise.
    Use for cells where the template has its own default: writing 0 would cause DIV/0
    in formula cells that divide by these inputs (e.g. L11=L10/L7, L8=(L7-K7)/K7).
    """
    if val is None or str(val).strip() in ('', 'null', 'None'):
        return None
    try:
        v = float(val)
        return None if v != v else v  # NaN → None
    except (TypeError, ValueError):
        return None


def generate_excel(all_inputs: dict, results: dict, detected_fy_years: list) -> io.BytesIO:
    """
    Load the Prebid V31 template, fill ~70 INPUT cells from all_inputs,
    leave all formula cells untouched, and return a BytesIO buffer.
    """
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb['Sheet1']
    ws_tf = wb['Transaction Fees']

    # ── Company Name ─────────────────────────────────────────────
    company_name = all_inputs.get('company_name') or 'Company'
    ws['A1'] = f'Project {company_name}'
    ws['C6'] = datetime.date.today()

    # ── Fiscal Year Labels (I4, J4, K4) ─────────────────────────
    fy_labels = []
    for i, fallback in enumerate(['FY1', 'FY2', 'FY3']):
        label = all_inputs.get(f'fy_year_{i+1}')
        if not label and len(detected_fy_years) > i:
            label = str(detected_fy_years[i])
        fy_labels.append(label or fallback)
    ws['I4'], ws['J4'], ws['K4'] = fy_labels

    # ── Sources (left side, cols C/D) ────────────────────────────
    ws['C7']  = _f(all_inputs.get('net_revenue_collateral'))
    ws['D7']  = _f(all_inputs.get('net_revenue_multiplier', 0.75))
    ws['C8']  = _f(all_inputs.get('inventory_collateral'))
    ws['D8']  = _f(all_inputs.get('inventory_multiplier', 0.70))
    ws['C9']  = _f(all_inputs.get('me_equipment_collateral'))
    ws['D9']  = _f(all_inputs.get('me_equipment_multiplier', 0.50))
    ws['C11'] = _f(all_inputs.get('building_land_collateral'))
    ws['D11'] = _f(all_inputs.get('building_land_multiplier', 0.50))
    # C12 = Existing Term Loans. Fallback chain (mirrors calculator.py):
    # 1. explicit existing_term_loans
    # 2. line_of_credit + current_lt_debt (balance sheet components)
    # 3. adj_ebitda_fy3 × leverage_multiple (or derived from reported + adjustments)
    _existing_tl = _fopt(all_inputs.get('existing_term_loans'))
    if _existing_tl is None:
        _loc  = _fopt(all_inputs.get('line_of_credit'))
        _cltd = _fopt(all_inputs.get('current_lt_debt'))
        if _loc is not None or _cltd is not None:
            _existing_tl = (_loc or 0.0) + (_cltd or 0.0)
            logger.info(
                f"  Excel C12: existing_term_loans null — "
                f"using LOC({_loc}) + current_lt_debt({_cltd}) = {_existing_tl}"
            )
    if _existing_tl is None or _existing_tl == 0.0:
        _adj3 = _fopt(all_inputs.get('adj_ebitda_fy3'))
        _lev  = _f(all_inputs.get('leverage_multiple', 3.5))
        # Mirror calculator.py: if adj_ebitda_fy3 null/0, derive from reported + adjustments
        if _adj3 is None or _adj3 == 0.0:
            _rep3  = _fopt(all_inputs.get('reported_ebitda_fy3'))
            _adjs3 = _fopt(all_inputs.get('adjustments_fy3'))
            if _rep3 is not None and _adjs3 is not None:
                _adj3 = _rep3 + _adjs3
        if _adj3 is not None and _adj3 > 0:
            _existing_tl = round(_adj3 * _lev, 2)
            logger.info(
                f"  Excel C12: adj_ebitda({_adj3:,.0f}) × leverage({_lev}) = {_existing_tl:,.2f}"
            )
    ws['C12'] = _existing_tl if (_existing_tl is not None and _existing_tl > 0) else 0.0
    ws['C14'] = _f(all_inputs.get('seller_note'))
    ws['C16'] = _f(all_inputs.get('earnout'))
    ws['C18'] = _f(all_inputs.get('equity_roll_from_seller'))

    # ── Uses (left side) ────────────────────────────────────────
    # C26 = EBITDA for valuation.
    # Fallback chain: adj_ebitda_fy3 → derive from fy3 P&L → adj_ebitda_fy2 → derive from fy2 → adj_ebitda_fy1
    def _ebitda_from_pl(n):
        gm  = _f(all_inputs.get(f'gross_margin_fy{n}'))
        sg  = _f(all_inputs.get(f'sga_fy{n}'))
        adj = _f(all_inputs.get(f'adjustments_fy{n}', 0))
        return round(gm - sg + adj, 2) if gm is not None else 0.0

    ebitda_fy3 = _f(all_inputs.get('adj_ebitda_fy3'))
    ebitda_fy2 = _f(all_inputs.get('adj_ebitda_fy2'))
    ebitda_fy1 = _f(all_inputs.get('adj_ebitda_fy1'))

    if ebitda_fy3 and ebitda_fy3 > 0:
        ws['C26'] = ebitda_fy3
    elif _ebitda_from_pl(3) > 0:
        ws['C26'] = _ebitda_from_pl(3)
    elif ebitda_fy2 and ebitda_fy2 > 0:
        ws['C26'] = ebitda_fy2
        logger.info("  Excel C26: using adj_ebitda_fy2 (fy3 unavailable)")
    elif _ebitda_from_pl(2) > 0:
        ws['C26'] = _ebitda_from_pl(2)
        logger.info("  Excel C26: derived from fy2 P&L (fy3 unavailable)")
    elif ebitda_fy1 and ebitda_fy1 > 0:
        ws['C26'] = ebitda_fy1
        logger.info("  Excel C26: using adj_ebitda_fy1 (fy3+fy2 unavailable)")
    else:
        ws['C26'] = 0.0
        logger.warning("  Excel C26: all EBITDA sources null — C26 set to 0")
    ws['C27'] = _f(all_inputs.get('acquisition_multiple', 5.0))   # EV/EBITDA deal multiple (Jira: default 5x)
    ws['C28'] = _f(all_inputs.get('pct_acquired', 1.0))            # % of company acquired
    ws['C40'] = 6  # Exit Multiple (fixed)

    # ── Rate Parameters ─────────────────────────────────────────
    # H26 = effective tax rate in the Prebid V31 template.
    # capex_pct_availability is an internal calculator parameter — not written to Excel.
    tax = _fopt(all_inputs.get('effective_tax_rate'))
    if tax is not None and 0.05 <= tax <= 0.45:
        ws['H26'] = tax
        logger.info(f"  Excel H26: effective_tax_rate = {tax:.0%}")
    # (if not extracted, template default 30% remains — acceptable fallback)
    ws['H39'] = _f(all_inputs.get('depreciation_rate', 0.045))
    ws['H40'] = _f(all_inputs.get('mgmt_ltip_rate', 0.055))
    ws['H41'] = _f(all_inputs.get('atar_ownership_rate', 0.05))
    ws['H45'] = _f(all_inputs.get('return_of_equity_years', 3))
    ws['H46'] = _f(all_inputs.get('atar_repayment_years', 4))
    ws['A44'] = _f(all_inputs.get('lp_pct', 0.03))
    ws['A47'] = _f(all_inputs.get('preferred_pct', 0.05))
    ws['A52'] = _f(all_inputs.get('fccr_rate', 0.08))
    ws['A54'] = _f(all_inputs.get('remaining_cash_pct', 0.75))

    # ── Historical Data (I=FY1, J=FY2, K=FY3) ──────────────────
    # Use _fopt: skip write when value absent — K7 is a divisor in L8=(L7-K7)/K7
    for col, n in [('I', 1), ('J', 2), ('K', 3)]:
        for row, key in [(7, f'revenue_fy{n}'), (10, f'gross_margin_fy{n}'),
                         (13, f'sga_fy{n}'), (17, f'adjustments_fy{n}'),
                         (22, f'interest_expense_fy{n}')]:
            v = _fopt(all_inputs.get(key))
            if v is not None:
                ws[f'{col}{row}'] = v

    # ── Projections (L=Y1, M=Y2, N=Y3, O=Y4, P=Y5) ────────────
    # Use _fopt: skip write when value absent — L7/M7… are divisors in GM%/EBITDA% formulas
    qofe_val = _f(all_inputs.get('qof_e_diligence'))
    mgmt_fee = -abs(qofe_val) if qofe_val else 0

    for col, n in [('L', 1), ('M', 2), ('N', 3), ('O', 4), ('P', 5)]:
        for row, key in [(7, f'revenue_y{n}'), (10, f'gross_margin_y{n}'),
                         (13, f'sga_y{n}'), (17, f'adjustments_y{n}'),
                         (30, f'capex_y{n}'),      # CapEx projection row
                         (32, f'term_loan_y{n}')]:
            v = _fopt(all_inputs.get(key))
            if v is not None:
                ws[f'{col}{row}'] = v
        ws[f'{col}53'] = mgmt_fee
    # NOTE: Do NOT fill L21-P21 — row 22 formulas handle projection interest

    # ── Scenario Modeling: unlock input cells, lock formula cells ───────────
    # Build the full list of user-editable input cells
    _input_cells = [
        'A1', 'C6', 'I4', 'J4', 'K4',                          # labels / date
        'C7', 'D7', 'C8', 'D8', 'C9', 'D9',                    # sources (rows 7-9)
        'C11', 'D11', 'C12', 'C14', 'C16', 'C18',              # sources (rows 11-18)
        'C26', 'C27', 'C28', 'C40',                             # uses / deal terms
        'H26', 'H39', 'H40', 'H41', 'H45', 'H46',              # rate parameters
        'A44', 'A47', 'A52', 'A54',                             # LP / preferred / FCCR
    ]
    for _col in ['I', 'J', 'K']:                                # historical P&L
        _input_cells += [f'{_col}7', f'{_col}10', f'{_col}13', f'{_col}17', f'{_col}22']
    for _col in ['L', 'M', 'N', 'O', 'P']:                     # projection P&L
        _input_cells += [f'{_col}7', f'{_col}10', f'{_col}13', f'{_col}17', f'{_col}30', f'{_col}32', f'{_col}53']

    for _ref in _input_cells:
        ws[_ref].protection = Protection(locked=False)

    # Apply sheet protection — only unlocked cells above are editable in Excel
    ws.protection.sheet = True
    ws.protection.selectLockedCells = False   # user can still navigate/read formula cells

    # ── Transaction Fees Sheet ──────────────────────────────────
    ws_tf['D6']  = _f(all_inputs.get('debt_sourcing_rate', 0.0075))
    ws_tf['D8']  = _f(all_inputs.get('lawyers_rate', 0.0075))
    ws_tf['E10'] = _f(all_inputs.get('qof_e_diligence', 250))
    ws_tf['E11'] = _f(all_inputs.get('tax_fee', 125))
    ws_tf['E12'] = _f(all_inputs.get('rw_insurance', 50))
    ws_tf['E13'] = _f(all_inputs.get('atar_bonuses_senior', 75))
    ws_tf['E14'] = _f(all_inputs.get('atar_bonuses_junior', 300))
    ws_tf['E15'] = _f(all_inputs.get('project_other', 100))

    # Lock Transaction Fees tab entirely — structural helper sheet, not for user edits
    ws_tf.protection.sheet = True
    ws_tf.protection.selectLockedCells = False

    # ── Force auto-recalculation on open ────────────────────────
    wb.calculation.calcMode = 'auto'
    wb.calculation.fullCalcOnLoad = True

    # ── Save to buffer ──────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    logger.info(f"Excel export generated for '{company_name}'")
    return buf
